"""本地文件读写：JSON / CSV / Excel（Excel 需 openpyxl）。"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any


EXPORT_HEADERS = [
    "公司名称",
    "法人",
    "状态",
    "成立日期",
    "注册资本",
    "城市",
    "地址",
    "电话",
    "邮箱",
    "网址",
    "经营范围",
    "意向分",
    "动作",
    "分池",
    "命中规则",
    "命中关键词",
    "标签",
    "评分说明",
]


def load_records(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json(path)
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_excel(path)
    raise SystemExit(f"不支持的文件类型: {suffix}（请用 .json / .csv / .xlsx）")


def _load_json(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    if isinstance(data, dict):
        # 兼容接口结构: {"status":0,"data":{"rows":[...]}}
        rows = data.get("data", {}).get("rows") if isinstance(data.get("data"), dict) else None
        if isinstance(rows, list):
            return [x for x in rows if isinstance(x, dict)]
        if isinstance(data.get("rows"), list):
            return [x for x in data["rows"] if isinstance(x, dict)]
    raise SystemExit("JSON 格式无法识别：需要数组，或含 data.rows / rows 的对象")


def _load_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _load_excel(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("读取 Excel 需要安装 openpyxl：pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []

    records: list[dict[str, Any]] = []
    for row in rows_iter:
        item: dict[str, Any] = {}
        empty = True
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            if value not in (None, ""):
                empty = False
            item[header] = "" if value is None else value
        if not empty:
            records.append(item)
    return records


def save_excel(path: Path, rows: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise SystemExit("写出 Excel 需要安装 openpyxl：pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "企业数据"
    ws.append(EXPORT_HEADERS)

    for row in rows:
        ws.append([
            row.get("company_name", ""),
            row.get("legal_person", ""),
            row.get("company_status", ""),
            row.get("establish_date", ""),
            row.get("capital", ""),
            row.get("city", ""),
            row.get("address", ""),
            row.get("phone", ""),
            row.get("email", ""),
            row.get("website", ""),
            row.get("business_scope", ""),
            row.get("score", ""),
            row.get("action", ""),
            row.get("pool", ""),
            row.get("matched_rules", ""),
            row.get("matched_keywords", ""),
            row.get("tags", ""),
            row.get("reasons", ""),
        ])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def save_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = [
        "company_name",
        "legal_person",
        "company_status",
        "establish_date",
        "capital",
        "city",
        "address",
        "phone",
        "email",
        "website",
        "business_scope",
        "score",
        "action",
        "pool",
        "matched_rules",
        "matched_keywords",
        "tags",
        "reasons",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
